from openpyxl import load_workbook
import pandas as pd
from copy import copy

wb = load_workbook("main.xlsx")
ws = wb.active

align_center = copy(ws['B1'].alignment)
font_bold = copy(ws['B1'].font)
fill_blue = copy(ws['B1'].fill)
thin_border = copy(ws['B1'].border)
text = ws['B1'].value

ws['B2'].alignment = align_center
ws['B2'].font = font_bold
ws['B2'].fill = fill_blue
ws['B2'].border = thin_border
ws['B2'].value = text

sheet_shop = wb['리스트']
sheet_shop['A3'].value
sheet_shop.cell(row=3, column=1).value
sheet_shop.cell(3, 1).value


wb.save('format.xlsx')